# Excel Processor for Databricks
# This script processes Excel files containing customer and supplier data and extracts them into CSV files

from pyspark.sql import SparkSession
import os
import pandas as pd
import glob
from openpyxl.utils.exceptions import InvalidFileException
import datetime
import traceback
import csv
import logging
from io import StringIO
from pyspark.sql.types import StructType, StructField, StringType, DoubleType
import openpyxl

# Initialize Spark session
spark = SparkSession.builder.appName("ExcelDataProcessor").getOrCreate()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Configure paths for Databricks environment
input_directory = "dbfs:/mnt/source_data/Data_From_OneDrive/Data/S&P/lot6_clean/"
output_directory = "dbfs:/mnt/source_data/Data/s_and_p_csv/lot6aoutput"
log_directory = "dbfs:/mnt/source_data/Data/s_and_p_csv/lot6aoutput/logs"
debug_mode = False

# Create timestamp for this processing run
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# Check if directories exist and create if needed
def ensure_dir_exists(dir_path):
    try:
        dbutils.fs.ls(dir_path)
    except:
        dbutils.fs.mkdirs(dir_path)
        logger.info(f"Created directory: {dir_path}")

# Convert dbfs path to file system path
def dbfs_to_local(dbfs_path):
    if dbfs_path.startswith("dbfs:/"):
        return "/dbfs" + dbfs_path[5:]
    return dbfs_path

# Extract company name from filename
def extract_company_name(filename):
    # Remove file extension
    filename = os.path.splitext(filename)[0]
    
    # Split by underscore
    parts = filename.split('_')
    
    # Find the part that comes before _Customers_ or _Suppliers_
    company_name = ""
    for i, part in enumerate(parts):
        if part in ["Customers", "Suppliers"] and i > 0:
            company_name = parts[i-1]
            break
    
    # If not found, use a different approach
    if not company_name:
        # Try to extract from SPGlobal_CompanyName_...
        if filename.startswith("SPGlobal_"):
            # Remove SPGlobal_
            rest = filename[9:]
            # Find the part before the next underscore
            if "_" in rest:
                company_name = rest.split("_")[0]
            else:
                company_name = rest
    
    return company_name

# Extract data with pandas as a fallback method
def extract_data_with_pandas(file_path, file_type, company_name, debug=False):
    data_records = []
    
    try:
        # Try to read the Excel file with pandas
        excel_data = pd.ExcelFile(file_path)
        
        # First, read the entire sheet to identify section markers
        for sheet_name in excel_data.sheet_names:
            if debug:
                logger.debug(f"Processing sheet {sheet_name} with pandas")
            
            # Read the sheet without headers
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Find rows that contain section markers
            section_markers = {}
            for i, row in df_raw.iterrows():
                for j, cell in enumerate(row):
                    if isinstance(cell, str) and "disclosed" in cell.lower():
                        section_markers[i] = cell.strip()
                        if debug:
                            logger.debug(f"Found section marker at row {i}: {cell}")
                        break
            
            # Process each section
            for section_row, section_name in section_markers.items():
                # Check if this section has "No Data Available"
                has_no_data = False
                for i in range(section_row + 1, min(section_row + 5, len(df_raw))):
                    for j, cell in enumerate(df_raw.iloc[i]):
                        if isinstance(cell, str) and "no data available" in cell.lower():
                            has_no_data = True
                            if debug:
                                logger.debug(f"Section '{section_name}' has 'No Data Available'")
                            break
                    if has_no_data:
                        break
                
                # Skip this section if it has no data
                if has_no_data:
                    if debug:
                        logger.debug(f"Skipping section '{section_name}' because it has no data")
                    continue
                
                # Look for header row after the section marker
                header_row = None
                for i in range(section_row + 1, min(section_row + 10, len(df_raw))):
                    for j, cell in enumerate(df_raw.iloc[i]):
                        key_term = "customer name" if file_type == "customer" else "supplier name"
                        if isinstance(cell, str) and key_term in cell.lower():
                            header_row = i
                            if debug:
                                logger.debug(f"Found header row at {header_row} for section: {section_name}")
                            break
                    if header_row is not None:
                        break
                
                if header_row is not None:
                    # Find the next section marker (if any)
                    next_section_row = float('inf')
                    for row in section_markers.keys():
                        if row > header_row and row < next_section_row:
                            next_section_row = row
                    
                    # Read the data with the header row
                    try:
                        # Calculate number of rows to read
                        nrows = next_section_row - header_row - 1 if next_section_row < float('inf') else None
                        
                        df_section = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=nrows)
                        
                        # Clean the dataframe
                        df_section = df_section.dropna(how='all')
                        
                        # Add company name column
                        df_section["Company Name"] = company_name
                        
                        # Add disclosed information column from the section name
                        df_section["Disclosed Information"] = section_name
                        
                        # Try to extract description
                        description = ""
                        for i in range(section_row):
                            for j, cell in enumerate(df_raw.iloc[i]):
                                if isinstance(cell, str) and "KEY:" in cell:
                                    description = cell.strip()
                                    break
                            if description:
                                break
                        
                        # Add description column
                        df_section["Description"] = description
                        
                        # Convert to records
                        records = df_section.to_dict('records')
                        
                        # Filter out non-data rows
                        for record in records:
                            # Skip rows without key column data
                            key_field = "Customer Name" if file_type == "customer" else "Supplier Name"
                            if key_field not in record or record[key_field] == "":
                                continue
                            
                            # Skip rows that look like headers or section titles
                            key_value = str(record[key_field]).lower()
                            if "disclosed" in key_value or "name" in key_value or "no data" in key_value:
                                continue
                            
                            # Add the record
                            data_records.append(record)
                        
                        if debug and data_records:
                            logger.debug(f"Found {len(data_records)} records for section: {section_name}")
                    
                    except Exception as e:
                        if debug:
                            logger.debug(f"Error processing section {section_name}: {e}")
            
            # If no data found using section approach, try a more direct approach
            if not data_records:
                if debug:
                    logger.debug("No data found using section markers, trying alternative approach")
                
                # Try different header rows
                for header_row in range(0, 30):
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                        
                        # Check if this looks like a data table
                        key_column = "Customer Name" if file_type == "customer" else "Supplier Name"
                        
                        if key_column in df.columns:
                            if debug:
                                logger.debug(f"Found {key_column} column in header row {header_row}")
                            
                            # Check if there's a "No Data Available" right after the header
                            no_data_after_header = False
                            if header_row + 1 < len(df_raw):
                                for j, cell in enumerate(df_raw.iloc[header_row + 1]):
                                    if isinstance(cell, str) and "no data" in cell.lower():
                                        no_data_after_header = True
                                        if debug:
                                            logger.debug(f"Found 'No Data Available' after header at row {header_row}")
                                        break
                            
                            if no_data_after_header:
                                # Skip this header as it has no data
                                continue
                            
                            # Clean the dataframe
                            df = df.dropna(how='all')
                            
                            # Add company name column
                            df["Company Name"] = company_name
                            
                            # Try to find a section name
                            section_name = ""
                            for i in range(max(0, header_row - 5), header_row):
                                if i < len(df_raw):
                                    for j, cell in enumerate(df_raw.iloc[i]):
                                        if isinstance(cell, str) and "disclosed" in cell.lower():
                                            section_name = cell.strip()
                                            break
                                if section_name:
                                    break
                            
                            # Add disclosed information column
                            df["Disclosed Information"] = section_name
                            
                            # Try to extract description
                            description = ""
                            for i in range(header_row):
                                if i < len(df_raw):
                                    for j, cell in enumerate(df_raw.iloc[i]):
                                        if isinstance(cell, str) and "KEY:" in cell:
                                            description = cell.strip()
                                            break
                                if description:
                                    break
                            
                            # Add description column
                            df["Description"] = description
                            
                            # Convert to records
                            records = df.to_dict('records')
                            
                            # Filter out non-data rows
                            for record in records:
                                # First replace NaN with empty strings
                                for key, value in record.items():
                                    if pd.isna(value):
                                        record[key] = ""
                                
                                # Skip rows that look like headers or section titles
                                key_value = str(record[key_column]).lower()
                                if "disclosed" in key_value or "name" in key_value or "no data" in key_value:
                                    continue
                                
                                # Round specific numeric columns to 2 decimal places
                                numeric_columns = [
                                    "Supplier Expense ($M)", "Supplier Expense (%)", 
                                    "Min %", "Max %", "Min Value ($M)", "Max Value ($M)",
                                    "Customer Revenue ($M)", "Customer Revenue (%)"
                                ]
                                
                                for col in numeric_columns:
                                    if col in record and not pd.isna(record[col]) and isinstance(record[col], (int, float)):
                                        record[col] = round(record[col], 2)
                                
                                # Add the record
                                data_records.append(record)
                            
                            if data_records:
                                if debug:
                                    logger.debug(f"Found {len(data_records)} records with pandas (alternative approach)")
                                return data_records
                    
                    except Exception as e:
                        if debug:
                            logger.debug(f"Error with pandas at header row {header_row}: {e}")
                        continue
    
    except Exception as e:
        if debug:
            logger.debug(f"Error in extract_data_with_pandas: {e}")
        raise  # Re-raise the exception to be caught by the caller
    
    return data_records

def extract_data_from_excel(file_path, file_type, company_name, debug=False):
    data_records = []
    
    try:
        # Try reading with openpyxl first for more control over cell access
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if debug:
                logger.debug(f"Successfully loaded workbook with openpyxl")
                logger.debug(f"Sheet names: {workbook.sheetnames}")
            
            sheet = workbook.active
            
            # Find all section headers (Recently Disclosed or Prior and Not Recently Disclosed)
            section_markers = {}
            for row_idx in range(1, min(500, sheet.max_row + 1)):
                cell_value = sheet.cell(row=row_idx, column=1).value
                if cell_value and isinstance(cell_value, str):
                    cell_str = cell_value.strip()
                    if "recently disclosed" in cell_str.lower():
                        section_markers[row_idx] = cell_str
                        if debug:
                            logger.debug(f"Found section marker at row {row_idx}: {cell_str}")
            
            # Now look for header rows containing "Customer Name" or "Supplier Name"
            for section_row, section_name in section_markers.items():
                # Check if this section has "No Data Available"
                has_no_data = False
                for row_idx in range(section_row + 1, min(section_row + 5, sheet.max_row + 1)):
                    cell_value = sheet.cell(row=row_idx, column=1).value
                    if cell_value and isinstance(cell_value, str) and "no data available" in cell_value.lower():
                        has_no_data = True
                        if debug:
                            logger.debug(f"Section '{section_name}' has 'No Data Available'")
                        break
                
                # Skip this section if it has no data
                if has_no_data:
                    if debug:
                        logger.debug(f"Skipping section '{section_name}' because it has no data")
                    continue
                
                header_row_idx = None
                column_headers = {}
                
                # Check rows after each section marker
                for row_idx in range(section_row + 1, min(section_row + 10, sheet.max_row + 1)):
                    for col_idx in range(1, min(30, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        
                        if cell_value:
                            cell_str = str(cell_value).strip().lower()
                            
                            # Check if this is a header row
                            if (file_type == "customer" and "customer name" in cell_str) or \
                               (file_type == "supplier" and "supplier name" in cell_str):
                                header_row_idx = row_idx
                                
                                # Map column headers
                                for col in range(1, min(30, sheet.max_column + 1)):
                                    header_cell = sheet.cell(row=row_idx, column=col).value
                                    if header_cell:
                                        column_headers[col] = str(header_cell).strip()
                                
                                if debug:
                                    logger.debug(f"Found header row at {header_row_idx} for section: {section_name}")
                                    logger.debug(f"Column headers: {column_headers}")
                                
                                break
                    
                    if header_row_idx:
                        break
                
                # If header row was found, extract data for this section
                if header_row_idx:
                    # Find the next section marker (if any)
                    next_section_row = float('inf')
                    for row in section_markers.keys():
                        if row > header_row_idx and row < next_section_row:
                            next_section_row = row
                    
                    # Process data rows between this header and next section (or end of data)
                    for row_idx in range(header_row_idx + 1, min(next_section_row, sheet.max_row + 1)):
                        # Check if this row has data
                        row_has_data = False
                        key_value_present = False
                        
                        # First check if the key field (Customer/Supplier Name) is present
                        key_field_col = None
                        for col, header in column_headers.items():
                            key_field = "Customer Name" if file_type == "customer" else "Supplier Name"
                            if header == key_field:
                                key_field_col = col
                                break
                        
                        if key_field_col:
                            key_value = sheet.cell(row=row_idx, column=key_field_col).value
                            if key_value:
                                key_value_present = True
                                # Skip rows with no data available message
                                if isinstance(key_value, str) and "no data" in key_value.lower():
                                    if debug:
                                        logger.debug(f"Skipping row {row_idx} with 'No Data' in key field")
                                    continue
                        
                        # Check if the row has any data
                        for col in range(1, min(30, sheet.max_column + 1)):
                            if sheet.cell(row=row_idx, column=col).value:
                                row_has_data = True
                                break
                        
                        if not row_has_data or not key_value_present:
                            continue
                        
                        # Extract data from this row
                        row_data = {}
                        
                        # Set company name
                        row_data["Company Name"] = company_name
                        
                        # Set disclosed information from the section name
                        row_data["Disclosed Information"] = section_name
                        
                        # Extract description (often ticker symbol in format like OTCPK:LICT (MI KEY: 4137440; SPCIQ KEY: 402730))
                        description_value = None
                        
                        # Try to find description in the company description rows at the top
                        for row in range(1, section_row):
                            # Usually a description/ticker contains "KEY:" 
                            for col in [1, 2]:
                                cell_value = sheet.cell(row=row, column=col).value
                                if cell_value and isinstance(cell_value, str) and "KEY:" in cell_value:
                                    description_value = cell_value.strip()
                                    break
                            if description_value:
                                break
                        
                        if description_value:
                            row_data["Description"] = description_value
                        else:
                            row_data["Description"] = ""
                        
                        # Map other columns based on headers
                        for col, header in column_headers.items():
                            cell_value = sheet.cell(row=row_idx, column=col).value
                            
                            # Skip empty cells
                            if cell_value is None:
                                row_data[header] = ""
                                continue
                            
                            # Round specific numeric columns to 2 decimal places to match Excel display
                            numeric_columns = [
                                "Supplier Expense ($M)", "Supplier Expense (%)", 
                                "Min %", "Max %", "Min Value ($M)", "Max Value ($M)",
                                "Customer Revenue ($M)", "Customer Revenue (%)"
                            ]
                            
                            if header in numeric_columns and isinstance(cell_value, (int, float)):
                                # Round to 2 decimal places to match Excel display
                                cell_value = round(cell_value, 2)
                            # Clean string values
                            elif isinstance(cell_value, str):
                                cell_value = cell_value.strip()
                            
                            # Add to row data
                            row_data[header] = cell_value
                        
                        # Check if this is a valid data row (contains customer/supplier name)
                        key_field = "Customer Name" if file_type == "customer" else "Supplier Name"
                        
                        if key_field in row_data and row_data[key_field]:
                            # Skip rows that are just section headers or contain metadata
                            key_value = str(row_data[key_field]).lower()
                            if "disclosed" in key_value or "name" in key_value or "no data" in key_value:
                                continue
                            
                            # Add the record
                            data_records.append(row_data)
            
            # If we couldn't find data using section markers, try a more general approach
            if not data_records:
                # Similar general approach as in the original code...
                header_row_idx = None
                column_headers = {}
                
                for row_idx in range(1, min(50, sheet.max_row + 1)):
                    for col_idx in range(1, min(30, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        
                        if cell_value:
                            cell_str = str(cell_value).strip().lower()
                            
                            if (file_type == "customer" and "customer name" in cell_str) or \
                                (file_type == "supplier" and "supplier name" in cell_str):
                                # Found header row
                                header_row_idx = row_idx
                                
                                # Map headers and extract data as in the original code
                                # Map column headers
                                for col in range(1, min(30, sheet.max_column + 1)):
                                    header_cell = sheet.cell(row=row_idx, column=col).value
                                    if header_cell:
                                        column_headers[col] = str(header_cell).strip()
                                break
                    
                    if header_row_idx:
                        break
                
                # If header row was found, extract data
                if header_row_idx:
                    # Processing similar to the section-based extraction...
                    # Extract data using the header row
                    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                        # Extract row data similar to the original code
                        row_has_data = False
                        for col in range(1, min(30, sheet.max_column + 1)):
                            if sheet.cell(row=row_idx, column=col).value:
                                row_has_data = True
                                break
                        
                        if not row_has_data:
                            continue
                        
                        # Create and populate row_data dictionary
                        row_data = {}
                        row_data["Company Name"] = company_name
                        row_data["Disclosed Information"] = ""
                        row_data["Description"] = ""
                        
                        # Map other columns based on headers
                        for col, header in column_headers.items():
                            cell_value = sheet.cell(row=row_idx, column=col).value
                            
                            # Process cell value
                            if cell_value is None:
                                row_data[header] = ""
                            elif isinstance(cell_value, (int, float)) and header in [
                                "Supplier Expense ($M)", "Supplier Expense (%)", 
                                "Min %", "Max %", "Min Value ($M)", "Max Value ($M)",
                                "Customer Revenue ($M)", "Customer Revenue (%)"
                            ]:
                                row_data[header] = round(cell_value, 2)
                            elif isinstance(cell_value, str):
                                row_data[header] = cell_value.strip()
                            else:
                                row_data[header] = cell_value
                        
                        # Check if this is a valid data row
                        key_field = "Customer Name" if file_type == "customer" else "Supplier Name"
                        if key_field in row_data and row_data[key_field]:
                            # Skip rows that are just section headers or contain metadata
                            key_value = str(row_data[key_field]).lower()
                            if "disclosed" in key_value or "name" in key_value or "no data" in key_value:
                                continue
                            
                            # Add the record
                            data_records.append(row_data)
        
        except InvalidFileException as e:
            # Fall back to pandas if openpyxl fails
            data_records = extract_data_with_pandas(file_path, file_type, company_name, debug)
                
    except Exception as e:
        if debug:
            logger.debug(f"Error in extract_data_from_excel: {e}")
        raise  # Re-raise the exception to be caught by the caller
    
    return data_records

# Process a single Excel file
def process_single_file(file_path, output_directory, debug=False):
    filename = os.path.basename(file_path)
    
    try:
        # Extract company name from filename
        company_name = extract_company_name(filename)
        
        # Determine if file contains customer or supplier data based on filename
        is_customer_file = "_Customers_" in filename
        is_supplier_file = "_Suppliers_" in filename
        
        if not (is_customer_file or is_supplier_file):
            if debug:
                logger.debug(f"Cannot determine file type for {filename}, skipping...")
            
            return ("skip", filename, "", {
                "reason": "Cannot determine file type (missing '_Customers_' or '_Suppliers_' in filename)"
            })
        
        # Extract data from the file
        file_data = {"customer_data": [], "supplier_data": []}
        extraction_status = {"customer": False, "supplier": False}
        extraction_reason = {"customer": "", "supplier": ""}
        
        if is_customer_file:
            try:
                customer_data = extract_data_from_excel(file_path, "customer", company_name, debug)
                
                if customer_data:
                    file_data["customer_data"] = customer_data
                    extraction_status["customer"] = True
                else:
                    extraction_reason["customer"] = "No customer data found"
            except Exception as e:
                extraction_reason["customer"] = f"Error: {str(e)}"
        
        if is_supplier_file:
            try:
                supplier_data = extract_data_from_excel(file_path, "supplier", company_name, debug)
                
                if supplier_data:
                    file_data["supplier_data"] = supplier_data
                    extraction_status["supplier"] = True
                else:
                    extraction_reason["supplier"] = "No supplier data found"
            except Exception as e:
                extraction_reason["supplier"] = f"Error: {str(e)}"
        
        # Determine success or failure
        if (is_customer_file and extraction_status["customer"]) or (is_supplier_file and extraction_status["supplier"]):
            return ("success", filename, company_name, file_data)
        else:
            reasons = []
            if is_customer_file and not extraction_status["customer"]:
                reasons.append(f"Customer data: {extraction_reason['customer']}")
            if is_supplier_file and not extraction_status["supplier"]:
                reasons.append(f"Supplier data: {extraction_reason['supplier']}")
            
            return ("fail", filename, company_name, {
                "reasons": reasons
            })
            
    except Exception as e:
        error_traceback = traceback.format_exc()
        return ("fail", filename, extract_company_name(filename), {
            "reasons": [f"Unexpected error: {str(e)}"],
            "traceback": error_traceback
        })

# Main processing function
def process_excel_files():
    # Create timestamp for this processing run
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Initialize tracking lists
    successful_files = []
    failed_files = []
    skipped_files = []
    processing_summary = {
        "total_files": 0,
        "successful_files": 0,
        "failed_files": 0,
        "skipped_files": 0,
        "customer_records": 0,
        "supplier_records": 0
    }
    
    # Ensure output and log directories exist
    ensure_dir_exists(output_directory)
    ensure_dir_exists(log_directory)
    
    # Initialize lists to store combined data
    all_customer_data = []
    all_supplier_data = []
    
    # Define columns for the output CSVs
    customer_columns = [
        "Company Name", "Description", "Disclosed Information", 
        "Customer Name", "Supplier Name", "Relationship Type", 
        "Primary Industry", "Geography", "Start Date", "End Date", 
        "Customer Revenue ($M)", "Customer Revenue (%)", "Min %", "Max %",
        "Min Value ($M)", "Max Value ($M)", "Source"
    ]
    
    supplier_columns = [
        "Company Name", "Description", "Disclosed Information", 
        "Supplier Name", "Customer Name", "Relationship Type", 
        "Primary Industry", "Geography", "Start Date", "End Date", 
        "Supplier Expense ($M)", "Supplier Expense (%)", "Min %", "Max %",
        "Min Value ($M)", "Max Value ($M)", "Source"
    ]
    
    # Get list of Excel files
    excel_files = [f.path for f in dbutils.fs.ls(input_directory) if f.path.endswith('.xls') or f.path.endswith('.xlsx')]
    
    if not excel_files:
        logger.warning(f"No Excel files found in {input_directory}")
        return
    
    processing_summary["total_files"] = len(excel_files)
    logger.info(f"Found {len(excel_files)} Excel file(s)")
    
    # Process files one by one
    logger.info("Processing files...")
    for file_path in excel_files:
        local_path = dbfs_to_local(file_path)
        filename = os.path.basename(local_path)
        logger.info(f"Processing file: {filename}")
        
        result = process_single_file(local_path, output_directory, debug_mode)
        
        if result is None:
            continue
            
        status, filename, company_name, file_data = result
        
        if status == "success":
            customer_data = file_data.get("customer_data", [])
            supplier_data = file_data.get("supplier_data", [])
            
            if customer_data:
                all_customer_data.extend(customer_data)
                successful_files.append({
                    "filename": filename,
                    "company": company_name,
                    "customer_records": len(customer_data),
                    "supplier_records": 0
                })
                processing_summary["customer_records"] += len(customer_data)
            
            if supplier_data:
                all_supplier_data.extend(supplier_data)
                # Check if we already added this file to successful_files
                if customer_data:
                    # Update the existing entry
                    for entry in successful_files:
                        if entry["filename"] == filename:
                            entry["supplier_records"] = len(supplier_data)
                            break
                else:
                    successful_files.append({
                        "filename": filename,
                        "company": company_name,
                        "customer_records": 0,
                        "supplier_records": len(supplier_data)
                    })
                processing_summary["supplier_records"] += len(supplier_data)
            
            processing_summary["successful_files"] += 1
            
        elif status == "skip":
            skipped_files.append({
                "filename": filename,
                "reason": file_data.get("reason", "Unknown reason")
            })
            processing_summary["skipped_files"] += 1
            
        elif status == "fail":
            failed_files.append({
                "filename": filename,
                "company": company_name,
                "reasons": file_data.get("reasons", ["Unknown error"]),
                "traceback": file_data.get("traceback", "")
            })
            processing_summary["failed_files"] += 1
    
    if all_customer_data:
        # Convert records to DataFrame
        customer_df = pd.DataFrame(all_customer_data)
        
        # Ensure all required columns exist
        for col in customer_columns:
            if col not in customer_df.columns:
                customer_df[col] = ""
        
        # Reorder columns
        customer_df = customer_df[customer_columns]
        
        # Drop duplicates
        customer_df.drop_duplicates(inplace=True)
        
        # Fill NaN values with empty strings
        customer_df.fillna("", inplace=True)
        
        # For large datasets in Databricks, use this approach:
        # Convert to Spark DataFrame
        customer_spark_df = spark.createDataFrame(customer_df)
        
        # Force coalesce to 1 partition to get a single output file
        customer_spark_df = customer_spark_df.coalesce(1)
        
        # Write to CSV with a temporary directory name
        temp_output_path = f"{output_directory}/temp_customers_{timestamp}"
        customer_spark_df.write.mode("overwrite").option("header", "true").option("quoteAll", "true").csv(temp_output_path)
        
        # Get the part file path
        part_files = [f.path for f in dbutils.fs.ls(temp_output_path) if f.name.startswith("part-")]
        if part_files:
            # First remove any existing file
            try:
                dbutils.fs.rm(f"{output_directory}/customers.csv")
            except:
                pass
                
            # Rename the part file to the final CSV name
            dbutils.fs.mv(part_files[0], f"{output_directory}/customers.csv")
            
            # Clean up temporary directory
            dbutils.fs.rm(temp_output_path, recurse=True)
        
        logger.info(f"Saved {len(all_customer_data)} customer records to {output_directory}/customers.csv")

    # Write supplier data to CSV
    if all_supplier_data:
        # Convert records to DataFrame
        supplier_df = pd.DataFrame(all_supplier_data)
        
        # Ensure all required columns exist
        for col in supplier_columns:
            if col not in supplier_df.columns:
                supplier_df[col] = ""
        
        # Reorder columns
        supplier_df = supplier_df[supplier_columns]
        
        # Drop duplicates
        supplier_df.drop_duplicates(inplace=True)
        
        # Fill NaN values with empty strings
        supplier_df.fillna("", inplace=True)
        
        # For large datasets in Databricks, use this approach:
        # Convert to Spark DataFrame
        supplier_spark_df = spark.createDataFrame(supplier_df)
        
        # Force coalesce to 1 partition to get a single output file
        supplier_spark_df = supplier_spark_df.coalesce(1)
        
        # Write to CSV with a temporary directory name
        temp_output_path = f"{output_directory}/temp_suppliers_{timestamp}"
        supplier_spark_df.write.mode("overwrite").option("header", "true").option("quoteAll", "true").csv(temp_output_path)
        
        # Get the part file path
        part_files = [f.path for f in dbutils.fs.ls(temp_output_path) if f.name.startswith("part-")]
        if part_files:
            # First remove any existing file
            try:
                dbutils.fs.rm(f"{output_directory}/suppliers.csv")
            except:
                pass
                
            # Rename the part file to the final CSV name
            dbutils.fs.mv(part_files[0], f"{output_directory}/suppliers.csv")
            
            # Clean up temporary directory
            dbutils.fs.rm(temp_output_path, recurse=True)
        
        logger.info(f"Saved {len(all_supplier_data)} supplier records to {output_directory}/suppliers.csv")    
    
    # Generate processing summary and logs
    summary_log = f"{log_directory}/processing_summary_{timestamp}.txt"
    
    # Create a summary log as string
    summary_content = f"""===== Excel Data Extraction Summary =====
Run Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Input Directory: {input_directory}
Output Directory: {output_directory}

=== Processing Statistics ===
Total files found: {processing_summary['total_files']}
Successfully processed: {processing_summary['successful_files']}
Failed to process: {processing_summary['failed_files']}
Skipped (indeterminate type): {processing_summary['skipped_files']}
Total customer records extracted: {processing_summary['customer_records']}
Total supplier records extracted: {processing_summary['supplier_records']}

=== Successfully Processed Files ===
"""
    
    for file in successful_files:
        summary_content += f"- {file['filename']} (Company: {file['company']})\n"
        if file.get('customer_records', 0) > 0:
            summary_content += f"  Customer records: {file['customer_records']}\n"
        if file.get('supplier_records', 0) > 0:
            summary_content += f"  Supplier records: {file['supplier_records']}\n"
    
    summary_content += "\n=== Failed Files ===\n"
    for file in failed_files:
        summary_content += f"- {file['filename']} (Company: {file['company']})\n"
        summary_content += "  Reasons:\n"
        for reason in file['reasons']:
            summary_content += f"  - {reason}\n"
    
    summary_content += "\n=== Skipped Files ===\n"
    for file in skipped_files:
        summary_content += f"- {file['filename']}\n"
        summary_content += f"  Reason: {file.get('reason', 'Unknown')}\n"
    
    # Write summary log to DBFS
    dbutils.fs.put(summary_log, summary_content, overwrite=True)
    
    logger.info(f"\nSummary report written to: {summary_log}")
    
    # Also save detailed logs as CSV files
    # Convert tracking lists to DataFrames
    successful_df = pd.DataFrame([{
        'filename': f['filename'],
        'company': f['company'],
        'customer_records': f.get('customer_records', 0),
        'supplier_records': f.get('supplier_records', 0)
    } for f in successful_files])

    failed_df = pd.DataFrame([{
        'filename': f['filename'],
        'company': f['company'],
        'reasons': '; '.join(f['reasons'])
    } for f in failed_files])

    skipped_df = pd.DataFrame([{
        'filename': f['filename'],
        'reason': f.get('reason', 'Unknown')
    } for f in skipped_files])
    
    # Save to CSV
    if not successful_df.empty:
        successful_spark_df = spark.createDataFrame(successful_df)
        successful_spark_df.write.mode("overwrite").option("header", "true").csv(f"{log_directory}/successful_files_{timestamp}.csv")
    
    if not failed_df.empty:
        failed_spark_df = spark.createDataFrame(failed_df)
        failed_spark_df.write.mode("overwrite").option("header", "true").csv(f"{log_directory}/failed_files_{timestamp}.csv")
    
    if not skipped_df.empty:
        skipped_spark_df = spark.createDataFrame(skipped_df)
        skipped_spark_df.write.mode("overwrite").option("header", "true").csv(f"{log_directory}/skipped_files_{timestamp}.csv")
    
    # Print summary to console
    logger.info("\n===== Processing Summary =====")
    logger.info(f"Total files: {processing_summary['total_files']}")
    logger.info(f"Successfully processed: {processing_summary['successful_files']}")
    logger.info(f"Failed to process: {processing_summary['failed_files']}")
    logger.info(f"Skipped (indeterminate type): {processing_summary['skipped_files']}")
    logger.info(f"Total customer records extracted: {processing_summary['customer_records']}")
    logger.info(f"Total supplier records extracted: {processing_summary['supplier_records']}")
    
    return {
        "successful_files": successful_files,
        "failed_files": failed_files,
        "skipped_files": skipped_files,
        "summary": processing_summary
    }

# Execute the main function
try:
    start_time = datetime.datetime.now()
    logger.info(f"Processing started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    result = process_excel_files()
    
    end_time = datetime.datetime.now()
    duration = end_time - start_time
    logger.info(f"Processing completed at: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Total duration: {duration}")
    
    if result:
        logger.info(f"Successfully processed {result['summary']['successful_files']} of {result['summary']['total_files']} files")
        logger.info(f"Extracted {result['summary']['customer_records']} customer records and {result['summary']['supplier_records']} supplier records")
    
    print("\nProcessing complete!")
    print(f"Duration: {duration}")
    
except Exception as e:
    logger.error(f"Fatal error during processing: {e}")
    logger.error(traceback.format_exc())
    print(f"\nError occurred: {e}")